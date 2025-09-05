import os
import csv
import base64
import json
from gc_forms_api_client import GCFormsApiClient
from form_submission_decrypter import FormSubmissionDecrypter
from form_submission_integrity_verifier import FormSubmissionVerifier
from access_token_generator import AccessTokenGenerator
from data_structures import PrivateApiKey, FormSubmission
from types import SimpleNamespace
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

IDENTITY_PROVIDER_URL = "https://auth.forms-formulaires.alpha.canada.ca"
PROJECT_IDENTIFIER = "284778202772022819"
GCFORMS_API_URL = "https://api.forms-formulaires.alpha.canada.ca"
EXCEL_FILE = "python_decryption/ID_to_Question_DRS_Intake.xlsx"

def load_private_api_key() -> PrivateApiKey:
    try:
        encoded_key = os.environ.get("PRIVATE_KEY_B64")
        if not encoded_key:
            raise Exception("Environment variable PRIVATE_KEY_B64 not found")

        decoded_bytes = base64.b64decode(encoded_key)
        file_as_json_object = json.loads(decoded_bytes.decode("utf-8"))
        return PrivateApiKey.from_json(file_as_json_object)
    except Exception as exception:
        raise Exception("Failed to load private API key from environment") from exception

def update_excel_with_submissions(submissions):
    wb = load_workbook(EXCEL_FILE)
    question_sheet = wb["Question To ID"]
    submission_sheet = wb["Submissions"]

    # Build question ID to question text mapping
    question_map = {}
    for row in question_sheet.iter_rows(min_row=2, values_only=True):
        qid, question = row
        question_map[str(qid)] = question

    # Prepare header row
    headers = ["Submitted At"] + [question_map[qid] for qid in question_map]
    if submission_sheet.max_row == 1 and submission_sheet.max_column == 1 and submission_sheet.cell(1, 1).value is None:
        submission_sheet.append(headers)

    # Append each submission
    for submission in submissions:
        answers = json.loads(submission["answers"])
        submitted_at = submission["submitted_at"]
        row = [submitted_at]
        for qid in question_map:
            row.append(answers[str(qid)])
        submission_sheet.append(row)

    # Define table range
    end_col = submission_sheet.max_column
    end_row = submission_sheet.max_row
    table_ref = f"A1:{chr(64 + end_col)}{end_row}"
    table = Table(displayName="SubmissionsTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    submission_sheet.add_table(table)

    wb.save(EXCEL_FILE)

def main():
    private_api_key = load_private_api_key()
    responses_for_file = ""
    print("\nGenerating access token...")

    access_token = AccessTokenGenerator.generate(
        IDENTITY_PROVIDER_URL, PROJECT_IDENTIFIER, private_api_key
    )

    api_client = GCFormsApiClient(
        private_api_key.form_id, GCFORMS_API_URL, access_token
    )

    print("\nRetrieving form template...\n")
    form_template = api_client.get_form_template()

    print("\nRetrieving new form submissions...")
    new_form_submissions = api_client.get_new_form_submissions()

    collected_submissions = []

    if len(new_form_submissions) > 0:
        print("\nNew form submissions:")
        print(", ".join(x.name for x in new_form_submissions))

        print("\nRetrieving, decrypting and confirming form submissions...")

        for new_form_submission in new_form_submissions:
            print(f"\nProcessing {new_form_submission.name}...\n")

            print("Retrieving encrypted submission...")
            encrypted_submission = api_client.get_form_submission(
                new_form_submission.name
            )

            print("\nDecrypting submission...")
            decrypted_form_submission = FormSubmissionDecrypter.decrypt(
                encrypted_submission, private_api_key
            )

            form_submission = FormSubmission.from_json(
                json.loads(decrypted_form_submission)
            )

            print("\nVerifying submission integrity...")
            integrity_verification_result = FormSubmissionVerifier.verify_integrity(
                form_submission.answers, form_submission.checksum
            )

            print(
                f"\nIntegrity verification result: {'OK' if integrity_verification_result else 'INVALID'}"
            )

            print("\nConfirming submission...")
            api_client.confirm_form_submission(
                new_form_submission.name, form_submission.confirmation_code
            )

            print("\nSubmission confirmed")

            print(form_submission)

            collected_submissions.append({
                "answers": form_submission.answers,
                "submitted_at": form_submission.created_at
            })
    else:
        print("\nCould not find any new form submission!")

    if collected_submissions:
        update_excel_with_submissions(collected_submissions)

if __name__ == "__main__":
    main()




